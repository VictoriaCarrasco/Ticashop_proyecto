from datetime import date, datetime
from decimal import Decimal

from django.contrib import messages
from django.db.models import Count, Q, Sum
from django.http import HttpResponse, HttpResponseNotAllowed
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from .forms import EmpleadoForm, SolicitudVacacionalForm
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.template.loader import render_to_string
from io import BytesIO
from django.template.loader import get_template
from xhtml2pdf import pisa
from .models import Liquidacion, Empleado, Venta, ComisionVenta
from django.contrib.auth import login, logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import make_password
from .decorators import validar_rol
from calendar import monthrange
from .models import ComisionVenta 
from django.contrib import messages


from .forms import EmpleadoForm
from .models import (
    Empleado,
    SolicitudVacacional,
    Liquidacion,
    ComisionVenta,
    RegistroAsistencia,
)


# ===========================
# Dashboard
# ===========================
@login_required
def dashboard(request):
    hoy = timezone.localdate()
    total_empleados = Empleado.objects.count()
    presentes_hoy = RegistroAsistencia.objects.filter(fecha=hoy, estado="PRESENTE").count()
    vacaciones_pendientes = SolicitudVacacional.objects.filter(estado="PENDIENTE").count()
    liquidaciones_pendientes = Liquidacion.objects.filter(estado="PENDIENTE_FIRMA").count()
    comisiones_pendientes = ComisionVenta.objects.filter(estado="PENDIENTE").count()

    ctx = dict(
        total_empleados=total_empleados,
        presentes_hoy=presentes_hoy,
        vacaciones_pendientes=vacaciones_pendientes,
        liquidaciones_pendientes=liquidaciones_pendientes,
        comisiones_pendientes=comisiones_pendientes,
    )
    return render(request, "dashboard.html", ctx)


# ===========================
# Vacaciones
# ===========================
@login_required
def vacaciones_list(request):
    empleado_id = request.session.get("empleado_id")
    try:
        empleado = Empleado.objects.get(id=empleado_id)
    except Empleado.DoesNotExist:
        messages.error(request, "No se encontró tu información de empleado.")
        return redirect("dashboard")
    
    # Admin y RRHH ven todas, el resto solo las suyas
    if empleado.rol in ["ADMIN", "RRHH"]:
        qs = (
            SolicitudVacacional.objects
            .select_related("empleado")
            .order_by("-fecha_inicio", "-id")
        )
    else:
        qs = (
            SolicitudVacacional.objects
            .filter(empleado=empleado)
            .order_by("-fecha_inicio", "-id")
        )
    
    return render(request, "vacaciones_list.html", {"solicitudes": qs})


@login_required
@validar_rol(["ADMIN", "RRHH"])
def vacaciones_restablecer(request):
    """
    Restablece el saldo de vacaciones de todos los empleados a 15 días.
    Solo accesible para ADMIN y RRHH.
    """
    empleados = Empleado.objects.all()
    contador = 0
    
    for empleado in empleados:
        empleado.saldo_vacaciones_dias = 15
        empleado.save()
        contador += 1
    
    messages.success(request, f"Se restablecieron los días de vacaciones para {contador} empleados.")
    return redirect("vacaciones_list")


@login_required
def vacaciones_solicitar(request):
    empleado_id = request.session.get("empleado_id")
    try:
        empleado = Empleado.objects.get(id=empleado_id)
    except Empleado.DoesNotExist:
        messages.error(request, "No se encontró tu información de empleado.")
        return redirect("dashboard")
    
    if request.method == "POST":
        form = SolicitudVacacionalForm(request.POST)
        if form.is_valid():
            dias_solicitados = form.cleaned_data['dias']
            
            # Validar que tenga saldo suficiente
            if dias_solicitados > empleado.saldo_vacaciones_dias:
                messages.error(request, f"No tienes suficiente saldo. Disponible: {empleado.saldo_vacaciones_dias} días.")
                return render(request, "vacaciones_solicitar.html", {
                    'form': form,
                    'empleado': empleado,
                    'saldo_disponible': empleado.saldo_vacaciones_dias,
                })
            
            # Crear la solicitud
            solicitud = form.save(commit=False)
            solicitud.empleado = empleado
            solicitud.estado = "PENDIENTE"
            solicitud.dias = dias_solicitados
            solicitud.save()
            
            # Descontar del saldo del empleado
            empleado.saldo_vacaciones_dias -= dias_solicitados
            empleado.save()
            
            messages.success(request, f"Solicitud de {solicitud.dias} días enviada. Saldo restante: {empleado.saldo_vacaciones_dias} días.")
            return redirect("vacaciones_list")
        else:
            messages.error(request, "Por favor corrige los errores en el formulario.")
    else:
        form = SolicitudVacacionalForm()
    
    context = {
        'form': form,
        'empleado': empleado,
        'saldo_disponible': empleado.saldo_vacaciones_dias,
    }
    return render(request, "vacaciones_solicitar.html", context)


def vacaciones_export(request):
    # Crear el libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Solicitudes de Vacaciones"
    
    # Encabezados con estilo
    headers = ['Empleado', 'Fecha Inicio', 'Fecha Fin', 'Días', 'Estado']
    ws.append(headers)
    
    # Estilo para los encabezados
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Obtener datos
    solicitudes = SolicitudVacacional.objects.select_related("empleado").order_by("-fecha_inicio")
    
    for sol in solicitudes:
        ws.append([
            sol.empleado.nombre,
            sol.fecha_inicio.strftime("%d/%m/%Y"),
            sol.fecha_fin.strftime("%d/%m/%Y"),
            sol.dias,
            sol.get_estado_display()
        ])
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    
    # Centrar contenido
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="solicitudes_vacaciones.xlsx"'
    
    wb.save(response)
    return response


def vacaciones_nueva(request):
    # Placeholder: muestra un aviso (puedes crear un formulario si quieres)
    messages.info(request, "Formulario de nueva solicitud no implementado (demo).")
    return redirect("vacaciones_list")

def vacaciones_detalle(request, pk: int):
    v = get_object_or_404(SolicitudVacacional.objects.select_related("empleado"), pk=pk)
    # Reutilizamos la tabla y resaltamos la fila, o crea un template detalle si lo prefieres
    return render(request, "vacaciones_list.html", {"solicitudes": [v]})

@validar_rol(["ADMIN", "RRHH"])
def vacaciones_aprobar(request, pk: int):
    if request.method != "GET":
        return HttpResponseNotAllowed(["GET"])
    v = get_object_or_404(SolicitudVacacional, pk=pk)
    v.estado = "APROBADA"
    v.save(update_fields=["estado"])
    messages.success(request, f"Solicitud #{v.pk} aprobada.")
    return redirect("vacaciones_list")

@validar_rol(["ADMIN", "RRHH"])
def vacaciones_rechazar(request, pk: int):
    if request.method != "GET":
        return HttpResponseNotAllowed(["GET"])
    v = get_object_or_404(SolicitudVacacional, pk=pk)
    
    # Solo devolver días si estaba PENDIENTE
    if v.estado == "PENDIENTE":
        empleado = v.empleado
        empleado.saldo_vacaciones_dias += v.dias
        empleado.save()
    
    v.estado = "RECHAZADA"
    v.save(update_fields=["estado"])
    messages.success(request, f"Solicitud #{v.pk} rechazada. Días devueltos al saldo del empleado.")
    return redirect("vacaciones_list")



# ===========================
# Liquidaciones
# ===========================


def render_to_pdf(template_src, context_dict=None):
    """
    Renderiza una plantilla HTML a PDF usando xhtml2pdf.
    Devuelve bytes o None si hay error.
    """
    if context_dict is None:
        context_dict = {}

    template = get_template(template_src)
    html = template.render(context_dict)

    result = BytesIO()
    pdf = pisa.pisaDocument(
        BytesIO(html.encode("UTF-8")),
        dest=result,
        encoding="UTF-8"
    )

    if pdf.err:
        # Si hay error, devolvemos None
        return None

    return result.getvalue()


@login_required
def liquidaciones_list(request):
    """
    - Empleado con rol NOMINA: ve TODAS las liquidaciones.
    - Cualquier otro rol: solo ve SUS propias liquidaciones.
    """
    empleado_id = request.session.get("empleado_id")
    if not empleado_id:
        messages.error(request, "No se encontró el empleado asociado a tu sesión.")
        return redirect("dashboard")

    empleado = get_object_or_404(Empleado, id=empleado_id)

    if empleado.rol == "NOMINA":
        # Nómina ve todo
        qs = (
            Liquidacion.objects
            .select_related("empleado")
            .order_by("-periodo", "-id")
        )
    else:
        # Cualquier otro rol solo ve sus propias liquidaciones
        qs = (
            Liquidacion.objects
            .filter(empleado=empleado)
            .select_related("empleado")
            .order_by("-periodo", "-id")
        )

    return render(request, "liquidaciones_list.html", {"liquidaciones": qs})


@login_required
def liquidaciones_generar(request):
    """
    SOLO rol NOMINA puede generar/actualizar liquidaciones.
    """
    empleado_id = request.session.get("empleado_id")
    if not empleado_id:
        messages.error(request, "No se encontró el empleado asociado a tu sesión.")
        return redirect("dashboard")

    empleado = get_object_or_404(Empleado, id=empleado_id)

    if empleado.rol != "NOMINA":
        messages.error(request, "No tienes permisos para generar liquidaciones.")
        return redirect("liquidaciones_list")

    hoy = timezone.now()
    periodo = date(hoy.year, hoy.month, 1)

    empleados = Empleado.objects.all()
    creadas = 0

    for e in empleados:
        sueldo_base = Decimal(e.sueldo_fijo)

        comision_reg = ComisionVenta.objects.filter(
            empleado=e,
            periodo__year=periodo.year,
            periodo__month=periodo.month,
        ).order_by("-id").first()

        comision = comision_reg.comision if comision_reg else Decimal("0")

        total_days_in_month = monthrange(periodo.year, periodo.month)[1]

        ausencias = RegistroAsistencia.objects.filter(
            empleado=e,
            fecha__year=periodo.year,
            fecha__month=periodo.month,
            estado="AUSENTE"
        ).count()

        sueldo_diario = sueldo_base / Decimal(total_days_in_month)
        descuento_ausencias = sueldo_diario * Decimal(ausencias)
        monto_final = sueldo_base - descuento_ausencias

        # 👇 NO usar update_or_create, para no pisar el estado
        try:
            obj = Liquidacion.objects.get(empleado=e, periodo=periodo)
            # actualizamos montos, comisiones…
            obj.monto_total = monto_final
            obj.comisiones = comision

            # solo si NO está firmada, la dejamos pendiente
            if obj.estado != "FIRMADA":
                obj.estado = "PENDIENTE_FIRMA"

            obj.save()
        except Liquidacion.DoesNotExist:
            Liquidacion.objects.create(
                empleado=e,
                periodo=periodo,
                monto_total=monto_final,
                comisiones=comision,
                estado="PENDIENTE_FIRMA",
            )
            creadas += 1

    messages.success(
        request,
        f"Se generaron/actualizaron {creadas} liquidaciones con descuento por ausencias."
    )
    return redirect("liquidaciones_list")



@login_required
def liquidacion_firmar(request, pk: int):
    """
    Solo el dueño de la liquidación puede firmarla.
    """
    liquidacion = get_object_or_404(Liquidacion, pk=pk)

    empleado_id = request.session.get("empleado_id")
    if not empleado_id or liquidacion.empleado_id != empleado_id:
        messages.error(request, "No puedes firmar esta liquidación.")
        return redirect("liquidaciones_list")

    if request.method == "POST":
        firma_data = request.POST.get("firma")
        if not firma_data:
            messages.error(request, "No se recibió la firma.")
            return redirect("liquidacion_firmar", pk=pk)

        import base64
        from django.core.files.base import ContentFile

        format, imgstr = firma_data.split(";base64,")
        ext = format.split("/")[-1]
        firma_file = ContentFile(
            base64.b64decode(imgstr),
            name=f"firma_{liquidacion.id}.{ext}"
        )

        liquidacion.firma_empleado = firma_file
        liquidacion.fecha_firma = timezone.now()
        liquidacion.estado = "FIRMADA"
        liquidacion.save()

        messages.success(request, "Liquidación firmada correctamente.")
        return redirect("liquidaciones_list")

    return render(request, "liquidacion_firmar.html", {"liquidacion": liquidacion})


@login_required
def liquidaciones_detalle(request, pk: int):
    """
    - Rol NOMINA: puede ver cualquier liquidación.
    - Cualquier otro rol: solo puede ver su propia liquidación.
    """
    liquidacion = get_object_or_404(
        Liquidacion.objects.select_related("empleado"),
        pk=pk
    )

    empleado_id = request.session.get("empleado_id")
    if not empleado_id:
        messages.error(request, "No se encontró el empleado asociado a tu sesión.")
        return redirect("liquidaciones_list")

    empleado = get_object_or_404(Empleado, id=empleado_id)

    if empleado.rol != "NOMINA" and liquidacion.empleado_id != empleado.id:
        messages.error(request, "No puedes ver esta liquidación.")
        return redirect("liquidaciones_list")

    return render(
        request,
        "liquidaciones_list.html",
        {"liquidaciones": [liquidacion]}
    )


def get_descuento_ausencias(liquidacion):
    periodo = liquidacion.periodo
    empleado = liquidacion.empleado
    SUELDO_BASE_DEMO = Decimal("1000000.00")
    total_days_in_month = monthrange(periodo.year, periodo.month)[1]
    ausencias = RegistroAsistencia.objects.filter(
        empleado=empleado,
        fecha__year=periodo.year,
        fecha__month=periodo.month,
        estado="AUSENTE"
    ).count()
    sueldo_diario = SUELDO_BASE_DEMO / Decimal(total_days_in_month)
    return sueldo_diario * Decimal(ausencias)


@login_required
def liquidacion_pdf(request, pk: int):
    liquidacion = get_object_or_404(
        Liquidacion.objects.select_related("empleado"),
        pk=pk
    )

    # 👇 Ahora la comisión se obtiene desde ComisionVenta
    comisiones = get_comision_mes(liquidacion)

    sueldo_base = liquidacion.monto_total
    descuento_ausencias = get_descuento_ausencias(liquidacion)

    GRAT_RATE = Decimal("0.25")
    gratificacion = sueldo_base * GRAT_RATE
    imponible = sueldo_base + gratificacion + comisiones

    AFP_RATE = Decimal("0.10")
    SALUD_RATE = Decimal("0.07")
    SEGURO_RATE = Decimal("0.006")

    afp = imponible * AFP_RATE
    salud = imponible * SALUD_RATE
    seguro = imponible * SEGURO_RATE
    impuesto = Decimal("0")

    total_descuentos = afp + salud + seguro + impuesto + descuento_ausencias
    liquido = imponible - total_descuentos

    # 👇 opcional: sincronizar el campo en la BD por si quedó en 0
    if liquidacion.comisiones != comisiones:
        liquidacion.comisiones = comisiones
        liquidacion.save(update_fields=["comisiones"])

    context = {
        "liquidacion": liquidacion,
        "sueldo_base": sueldo_base,
        "gratificacion": gratificacion,
        "comisiones": comisiones,          # <-- comisión real
        "imponible": imponible,
        "afp": afp,
        "salud": salud,
        "seguro": seguro,
        "impuesto": impuesto,
        "total_descuentos": total_descuentos,
        "liquido": liquido,
        "descuento_ausencias": descuento_ausencias,
        "firma_empleado": liquidacion.firma_empleado,
        "fecha_firma": liquidacion.fecha_firma,
    }

    pdf_bytes = render_to_pdf("liquidacion_pdf.html", context)
    if pdf_bytes is None:
        return HttpResponse("Error al generar el PDF", status=500)

    filename = f"liquidacion_{liquidacion.empleado.rut}_{liquido:.0f}.pdf"
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename=\"{filename}\"'
    return response





# ===========================
# Comisiones
# ===========================

def get_comision_mes(liquidacion):
    """
    Busca la comisión del mes de esta liquidación en ComisionVenta.
    Si no hay registro, devuelve 0.
    """
    periodo = liquidacion.periodo
    empleado = liquidacion.empleado

    reg = ComisionVenta.objects.filter(
        empleado=empleado,
        periodo__year=periodo.year,
        periodo__month=periodo.month,
    ).order_by("-id").first()

    if reg and reg.comision is not None:
        return reg.comision

    return Decimal("0")


@validar_rol(["SUP_COM","ADMIN","GENERAL"])
def comisiones_list(request):
    qs = ComisionVenta.objects.select_related("empleado").order_by("-periodo", "-id")
    return render(request, "comisiones.html", {"comisiones": qs})

def comisiones_calcular(request):
    # DEMO: recalcula comisión = 1.5% de ventas_totales
    qs = ComisionVenta.objects.all()
    for c in qs:
        c.comision = (c.ventas_totales or Decimal("0")) * Decimal("0.015")
        c.estado = "CALCULADA"
        c.save(update_fields=["comision", "estado"])
    messages.success(request, "Comisiones recalculadas.")
    return redirect("comisiones_list")

def comisiones_export(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # Crear el libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Comisiones de Ventas"
    
    # Encabezados con estilo
    headers = ['Empleado', 'Período', 'Ventas Totales', 'Comisión', 'Estado']
    ws.append(headers)
    
    # Estilo para los encabezados
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Obtener datos
    comisiones = ComisionVenta.objects.select_related("empleado").order_by("-periodo")
    
    for com in comisiones:
        ws.append([
            com.empleado.nombre,
            com.periodo.strftime("%m/%Y"),
            f"${com.ventas_totales:,.0f}",
            f"${com.comision:,.0f}",
            com.estado
        ])
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # Centrar contenido
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="comisiones_ventas.xlsx"'
    
    wb.save(response)
    return response


def comisiones_detalle(request, pk: int):
    c = get_object_or_404(ComisionVenta.objects.select_related("empleado"), pk=pk)
    return render(request, "comisiones.html", {"comisiones": [c]})


# ===========================
# Asistencia
# ===========================
@validar_rol(["ADMIN", "NOMINA", "SUP_COM", "VENDEDOR", "RRHH"])
def asistencia_list(request):
    qs = RegistroAsistencia.objects.select_related("empleado").order_by("-fecha", "empleado__nombre")

    # filtros GET: desde, hasta, q (nombre o rut)
    desde = request.GET.get("desde")
    hasta = request.GET.get("hasta")
    q = request.GET.get("q")

    if desde:
        qs = qs.filter(fecha__gte=desde)
    if hasta:
        qs = qs.filter(fecha__lte=hasta)
    if q:
        qs = qs.filter(Q(empleado__nombre__icontains=q) | Q(empleado__rut__icontains=q))

    return render(request, "asistencia.html", {"registros": qs})

def asistencia_export(request):
    qs = RegistroAsistencia.objects.select_related("empleado").values_list(
        "empleado__nombre", "fecha", "hora_entrada", "hora_salida", "horas_trabajadas", "estado"
    )
    resp = HttpResponse(content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = 'attachment; filename="asistencia.csv"'
    resp.write("empleado,fecha,entrada,salida,horas,estado\n")
    for r in qs:
        resp.write(",".join([str(x) for x in r]) + "\n")
    return resp


# ===========================
# Usuarios (Empleados)
# ===========================
@login_required
def usuarios_list(request):
    if not request.user.is_staff:
        return redirect("dashboard")

    usuarios = User.objects.all().order_by("id")
    return render(request, "admin/usuarios_list.html", {"usuarios": usuarios})

def usuarios_new(request):
    if request.method == "POST":
        form = EmpleadoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Usuario creado.")
            return redirect("usuarios_list")
    else:
        form = EmpleadoForm()
    return render(request, "usuarios_form.html", {"form": form, "titulo": "Nuevo Usuario"})

def usuarios_view(request, pk: int):
    u = get_object_or_404(Empleado, pk=pk)
    # Para simplificar, reusamos la lista con solo ese usuario
    return render(request, "usuarios_list.html", {"usuarios": [u]})

def usuarios_edit(request, pk: int):
    u = get_object_or_404(Empleado, pk=pk)
    if request.method == "POST":
        form = EmpleadoForm(request.POST, instance=u)
        if form.is_valid():
            form.save()
            messages.success(request, "Usuario actualizado.")
            return redirect("usuarios_list")
    else:
        form = EmpleadoForm(instance=u)
    return render(request, "usuarios_form.html", {"form": form, "titulo": "Editar Usuario"})

def usuarios_delete(request, pk: int):
    u = get_object_or_404(Empleado, pk=pk)
    u.delete()
    messages.success(request, "Usuario eliminado.")
    return redirect("usuarios_list")

def usuarios_export(request):
    qs = Empleado.objects.values_list("nombre", "email", "rut", "rol", "departamento", "activo")
    resp = HttpResponse(content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = 'attachment; filename="usuarios.csv"'
    resp.write("nombre,email,rut,rol,departamento,activo\n")
    for r in qs:
        resp.write(",".join([str(x) for x in r]) + "\n")
    return resp



def login_empleado(request):
    """
    - Si viene con ?next=/admin/  → usa 'login_admin.html' (login admin estilizado)
    - Si no                       → usa 'login.html' (login empleados)
    """
    next_url = request.GET.get("next", "") or request.POST.get("next", "")
    if next_url is None:
        next_url = ""

    es_login_admin = next_url.startswith("/admin")
    template_name = "login_admin.html" if es_login_admin else "login.html"

    if request.method == "POST":
        rut = (request.POST.get("rut") or "").strip()
        password = request.POST.get("password") or ""

        if not rut or not password:
            messages.error(request, "Debes ingresar RUT y contraseña.")
            return render(request, template_name, {"next": next_url})

        try:
            empleado = Empleado.objects.get(rut=rut, activo=True)
        except Empleado.DoesNotExist:
            messages.error(request, "RUT no encontrado o empleado inactivo.")
            return render(request, template_name, {"next": next_url})

        try:
            user = User.objects.get(email=empleado.email, is_active=True)
        except User.DoesNotExist:
            messages.error(
                request,
                "No hay un usuario del sistema asociado al correo de este empleado."
            )
            return render(request, template_name, {"next": next_url})

        if not user.check_password(password):
            messages.error(request, "Contraseña incorrecta.")
            return render(request, template_name, {"next": next_url})

        login(request, user)
        request.session["empleado_id"] = empleado.id
        request.session["empleado_nombre"] = empleado.nombre
        request.session["empleado_rut"] = empleado.rut
        request.session["empleado_rol"] = empleado.rol

        if next_url:
            return redirect(next_url)
        return redirect("dashboard")

    return render(request, template_name, {"next": next_url})


def logout_empleado(request):
    logout(request)
    return redirect("login_empleado")





@login_required
def admin_home(request):
    # Solo staff puede ver el panel de administración
    if not request.user.is_staff:
        return redirect("dashboard")
    return render(request, "admin/admin_home.html")

@login_required
def admin_usuarios_list(request):
    if not request.user.is_staff:
        return redirect("dashboard")

    usuarios = User.objects.all().order_by("id")
    return render(request, "admin/usuarios_list.html", {"usuarios": usuarios})


@login_required
def empleados_list(request):
    if not request.user.is_staff:
        return redirect("dashboard")

    empleados = Empleado.objects.all().order_by("id")
    return render(
        request,
        "admin/empleados_list.html",
        {"empleados": empleados},
    )


@login_required
def empleado_edit(request, pk):
    if not request.user.is_staff:
        return redirect("dashboard")

    empleado = get_object_or_404(Empleado, pk=pk)

    if request.method == "POST":
        form = EmpleadoForm(request.POST, instance=empleado)
        if form.is_valid():
            form.save()
            messages.success(request, "Empleado actualizado correctamente.")
            return redirect("empleados_list")
    else:
        form = EmpleadoForm(instance=empleado)

    return render(
    request,
    "usuarios_form.html",         
    {"form": form, "titulo": "Editar empleado"},
)



@login_required
def empleado_toggle_activo(request, pk):
    if not request.user.is_staff:
        return redirect("dashboard")

    empleado = get_object_or_404(Empleado, pk=pk)
    empleado.activo = not empleado.activo
    empleado.save(update_fields=["activo"])

    estado = "activado" if empleado.activo else "desactivado"
    messages.success(request, f"Empleado {estado} correctamente.")
    return redirect("empleados_list")



SUELDOS_FIJOS_ROL = {
    "ADMIN": 1000000,
    "RRHH": 900000,
    "NOMINA": 850000,
    "JEFATURA": 800000,
    "SUP_COM": 750000,
    "TECNICO": 650000,
    "GENERAL": 500000,
}

@login_required
def crear_usuario_empleado(request):
    if not request.user.is_staff:
        return redirect("dashboard")

    if request.method == "POST":
        nombre = request.POST["nombre"]
        email = request.POST["email"]
        rut = request.POST["rut"]
        rol = request.POST["rol"]
        departamento = request.POST["departamento"]
        password = request.POST["password"]

        sueldo_fijo = SUELDOS_FIJOS_ROL.get(rol, 500000)  # Obtiene monto fijo por rol

        # 👇 NUEVO: definir si será staff o no
        es_staff = rol in ["ADMIN", "RRHH"]

        # Crear usuario Django
        user = User.objects.create(
            username=email.split("@")[0],
            email=email,
            password=make_password(password),
            is_active=True,
            is_staff=es_staff,   # 👈 ahora sí existe
        )

        # Crear empleado con SUELDO FIJO
        Empleado.objects.create(
            nombre=nombre,
            email=email,
            rut=rut,
            rol=rol,
            departamento=departamento,
            activo=True,
            sueldo_fijo=sueldo_fijo,
        )

        messages.success(request, "Usuario y empleado creados correctamente.")
        return redirect("admin_home")

    return render(request, "admin/crear_usuario_empleado.html")




@login_required
def usuario_create(request):
    # Reutilizamos el mismo formulario unificado
    return crear_usuario_empleado(request)
